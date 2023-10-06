import io
import zipfile, os
import utils
import xlrd
from utils import zip_dir
from openpyxl import load_workbook
from PyPDF2 import PdfReader



path_to_resources = utils.RESOURCES_PATH
path_to_zip = utils.TMP_PATH
number_files = len(os.listdir(path_to_resources))

def test_zipping_resources_files():
    if not os.path.exists('tmp'):
        os.mkdir('tmp')
    zip_dir(path_to_resources, f"{path_to_zip}/zip_resources_files.zip")
    with zipfile.ZipFile(f'{path_to_zip}/zip_resources_files.zip', mode='r') as zf:
        files_in_arch = len(zf.namelist())
        number_files_in_arch = files_in_arch
        assert number_files == number_files_in_arch

def test_reading_files_in_archive_and_checking():

    with zipfile.ZipFile(f'{path_to_zip}/zip_resources_files.zip', mode='r') as zf:
        #value of tmp/zip/xls file
        with zf.open('file_example_XLS_10.xls') as my_file_1:
            workbook = xlrd.open_workbook(file_contents=my_file_1.read())
            sheet = workbook.sheet_by_index(0)
            sheet_value_first_file_second_row_archive = sheet.row_values(1)

        #value of tmp/zip/xlsx file
        with zf.open('file_example_XLSX_50.xlsx') as my_file_2:
            workbook = load_workbook(filename=io.BytesIO(my_file_2.read()))
            sheet = workbook.active
            cell_value_second_file_xlsx_archive = sheet.cell(row=3, column=3).value

        # value of tmp/zip/zip file
        with zf.open('hello.zip') as inner_arch:
            my_file_3 = zipfile.ZipFile(io.BytesIO(inner_arch.read()))
            with my_file_3.open('Hello.txt') as txt_file:
                txt_file_archive_value = txt_file.read(10).decode()

        # value of tmp/zip/pdf file
        with zf.open('Python Testing with Pytest (Brian Okken).pdf') as my_file_4:
            pdf_file = PdfReader(io.BytesIO(my_file_4.read()))
            pdf_page_arch = pdf_file.pages[1].extract_text()

    #value of resources/xls file
    with xlrd.open_workbook(f'{path_to_resources}/file_example_XLS_10.xls') as workbook:
        sheet = workbook.sheet_by_index(0)
        sheet_value_first_row_xls = sheet.row_values(1)

    #value of resources/xlsx file
    work_book = load_workbook(f'{path_to_resources}/file_example_XLSX_50.xlsx')
    cell_3 = work_book.active
    cell_value_second_file_xlsx_resources = cell_3.cell(row=3, column=3).value
    work_book.close()

    #value of resources/zip file
    with zipfile.ZipFile(f"{path_to_resources}/hello.zip", mode='r') as res_zip:
        with res_zip.open('Hello.txt') as arch_txt_file:
            txt_file_res_value = arch_txt_file.read(10).decode()

    #value of resources/pdf file
    res_pdf_file = PdfReader(f"{path_to_resources}/Python Testing with Pytest (Brian Okken).pdf")
    pdf_page_res = res_pdf_file.pages[1].extract_text()

    assert sheet_value_first_row_xls == sheet_value_first_file_second_row_archive
    assert cell_value_second_file_xlsx_archive == cell_value_second_file_xlsx_resources
    assert txt_file_archive_value == txt_file_res_value
    assert pdf_page_arch == pdf_page_res



